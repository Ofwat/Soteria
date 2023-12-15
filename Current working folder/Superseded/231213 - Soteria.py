# -*- coding: utf-8 -*-

"""
Created on Wed Jan 18 14:16:11 2023

@author: h_jet
"""


from tkinter import Tk
from tkinter.filedialog import askdirectory
import os
import pandas as pd
import glob
import networkx as nx
import xlrd
from openpyxl import load_workbook
from pyxlsb import open_workbook
import time
import matplotlib.pyplot as plt
import pickle

pd.set_option('display.max_columns', 40)
pd.set_option('display.width', 2000)

start = time.time()

#  Select the parent folder for the model files.
path = askdirectory(title='Select Folder')  # This shows the dialog box and return the path.
print(path, "\n")

os.chdir(path)  # Make the parent folder the working directory

# Create list of all xls* files in root folder and subfolders and adds to xslxfiles if an "F_Inputs" or "F_Outputs" sheet is present
xlsxfiles = []
for file in glob.glob("**/*.xls*", recursive=True): # If recursive is true, the pattern “**” will match any files and zero or more directories, subdirectories and symbolic links to directories.
    filename, file_extension = os.path.splitext(file) 
    # check if files contain an F_Inputs or F_Outputs sheet
    try: 
        if file_extension == '.xls':
            filexl = xlrd.open_workbook(file)
            if 'F_Inputs' in filexl.sheet_names() or 'F_Outputs' in filexl.sheet_names():
                xlsxfiles.append(file)
    
        elif file_extension == '.xlsb':
            filexl = open_workbook(file)
            if 'F_Inputs' in filexl.sheets or 'F_Outputs' in filexl.sheets:
                xlsxfiles.append(file)
            
        else: 
            filexl = load_workbook(file, read_only = True)
            if 'F_Inputs' in filexl.sheetnames or 'F_Outputs' in filexl.sheetnames:
                xlsxfiles.append(file)
                    
    except:
        print("error with", file)


# Create lists to capture node attributes for all models

finputs_codes = []  # list of lists of F_Inputs boncodes
foutputs_codes = []  # list of lists of F_Outputs boncodes
finputs_dfs = []  # list of dataframes of F_Inputs sheets
foutputs_dfs = []  # list of dataframes of F_Outputs sheets
finputs_timestamps = [] # list of strings of F_Inputs timestamps
finputs_runs = []  # list of strings of the CLEAR_SHEET run IDs
foutputs_runs = []  # list of strings of the CLEAR_SHEET run IDs
finputs_reportids = []  # list of strings of the CLEAR_SHEET F_inputs run IDs
foutputs_reportids = []  # list of strings of the CLEAR_SHEET F_Outputs run IDs
finputs_companyids = []  # list of strings of the CLEAR_SHEET F_Inputs company IDs
foutputs_companyids = []  # list of strings of the CLEAR_SHEET F_Outputs company IDs
QA1s = []
QA2s = []
QA3s = []

inputsheet = 'F_Inputs'
outputsheet = 'F_Outputs'
clearsheet = 'CLEAR_sheet'
    
def scrape_model(xlfile_name, inputsheet, outputsheet, clearsheet):
    try:
        dfinput = pd.read_excel(file, sheet_name=inputsheet)  #  Create a dataframe from the F_Input sheet
        dfinput = dfinput.rename(columns=dfinput.iloc[0]).drop(dfinput.index[0])
        dfinput = dfinput.iloc[1:]
        finput = dfinput.Reference.tolist()  # Get a list of the BON codes?
        finput = [x for x in finput if str(x) !='nan']  # Drop NANs from the list of BON codes
    except ValueError:
        finput = []
     
    try:
        dfoutput = pd.read_excel(file, sheet_name=outputsheet)  # Create a dataframe from the F_Output sheet
        dfoutput = dfoutput.rename(columns=dfoutput.iloc[0]).drop(dfoutput.index[0])
        dfoutput = dfoutput.iloc[1:]
        foutput = dfoutput.Reference.tolist()   # Get the list of BON codes
        foutput = [x for x in finput if str(x) !='nan']  # Drop NANs from the list of BON codes
    except ValueError:
        foutput = []
        
    try:  # TODO: This name of this dataframe is same as above - better to give a different name?
        dfoutput = pd.read_excel(file, sheet_name=clearsheet)  #  Create a dataframe from the clearsheet
        
    except:
        finputid = ''
            
    return finput, foutput

for file in xlsxfiles:  # For each model file we are working with...
    finput, foutput = scrape_model(file, inputsheet, outputsheet, clearsheet)  # ...get the list of BON codes in the f_input and f_output sheets
    finputs_codes.append(finput)  # Add the list of bon codes. This is creating a list of lists.
    foutputs_codes.append(foutput)  # Add the list of bon codes. This is creating a list of lists.

# create digraph, nodes and edges
num_models = len(xlsxfiles)
G = nx.DiGraph()  # Create an empty digraph with no nodes and no edges.

# Add attribtues for each model called F_Inputs and F_Outputs
for i in range(num_models):  # For each of the model files...
    G.add_node(xlsxfiles[i],  # make each model a node...
               F_Inputs=finputs_codes[i],  # ...associate it with its list of F_Input BON codes and ...
               F_Outputs=foutputs_codes[i])  # ...associate it with its list of F_Input BON codes

for i in range(num_models):  # For each of the model files...
    for j in range(num_models):  # For each of the model files...
        data_trans = [k for k in finputs_codes[i] if k in foutputs_codes[j]]
        if data_trans:
            # Add edge between node j and i, associate with BON codes
            G.add_edge(xlsxfiles[j], xlsxfiles[i], Data_transfered=data_trans)


# Outputs  TODO: Could export this to excel or log file?
print("The model files used are: ", xlsxfiles, "\n")  # Print the files used
print("The nodes are: ", G.nodes, "\n")
#print("The notes data is as follows: ", G.nodes.data(), "\n")
print("The edges are: ", G.edges, "\n")
#print("The edges data is as follows: ", G.edges.data(), "\n")
print("The pandas edge list is: ", "\n", nx.to_pandas_edgelist(G), "\n")
print("The pandas adjacency matrix is: ", "\n", nx.to_pandas_adjacency(G), "\n")


# Draw network
nx.draw_networkx(G)
plt.show()

# Check for cycles
try:
    cycles = nx.find_cycle(G)
    print("Cycles found")
    print(*cycles, sep="/n")
except nx.exception.NetworkXNoCycle:
    print("No cycles found")
    
# Work out batch order
'''
def find_order(G):
    nodes = list(G.nodes)
    for node in nodes:
'''


end = time.time()
print("Time: ", round(end-start, 1))

