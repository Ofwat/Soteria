# -*- coding: utf-8 -*-

"""
Created on Wed Jan 18 14:16:11 2023

@author: h_jet
"""

#Next steps (not ordered)
# Add in scraping for QA1,2 and 3. - PARTIALLY DONE
# Functionalise the code for creating the newtork. Make this for singluar repetition
    # in case we need to add nodes and edges individually in the future
# Functionalise the code for creating the xlsxfiles list
    # also record the state of F_INputs, F_Outputs and CLEAR_SHEETs whilst testing the files
# convert all the lists into a dictionary of dictionaries instead - much more robust
# create run order creator - needs to decide what run order actually means
    # most accurate case would be to have run order per model (ie based on dependency)
    # can calculate concurrent run order which looks better but not as efficient
    # can also calculate linear run order but that is so innefficient it likely won't be used
# visualise. Possibly use pyvis. Come up with a way to procedurally generate node positions



from tkinter import Tk
from tkinter.filedialog import askdirectory
import os
import pandas as pd
import glob
import networkx as nx
import xlrd
from openpyxl import load_workbook
from pyxlsb import open_workbook
import time
import matplotlib.pyplot as plt
import pickle
import fnmatch 

pd.set_option('display.max_columns', 40)
pd.set_option('display.width', 2000)

start = time.time()

#  Select the parent folder for the model files.
path = askdirectory(title='Select Folder')  # This shows the dialog box and return the path.
print(path, "\n")

os.chdir(path)  # Make the parent folder the working directory

# Create list of all xls* files in root folder and subfolders and adds to xslxfiles if an "F_Inputs" or "F_Outputs" sheet is present
xlsxfiles = []


for file in glob.glob("**/*.xls*", recursive=True): # If recursive is true, the pattern “**” will match any files and zero or more directories, subdirectories and symbolic links to directories.
    filename, file_extension = os.path.splitext(file) 
    # check if files contain an F_Inputs or F_Outputs sheet
    try: 
        if file_extension == '.xls':
            filexl = xlrd.open_workbook(file)
            if 'F_Inputs' in filexl.sheet_names() or 'F_Outputs' in filexl.sheet_names():
                xlsxfiles.append(file)
    
        elif file_extension == '.xlsb':
            filexl = open_workbook(file)
            if 'F_Inputs' in filexl.sheets or 'F_Outputs' in filexl.sheets:
                xlsxfiles.append(file)
            
        else: 
            filexl = load_workbook(file, read_only = True)
            if 'F_Inputs' in filexl.sheetnames or 'F_Outputs' in filexl.sheetnames:
                xlsxfiles.append(file)
                    
    except:
        print("error with", file)


# Create lists to capture node attributes for all models

finputs_codes = []  # list of lists of F_Inputs boncodes
foutputs_codes = []  # list of lists of F_Outputs boncodes
finputs_dfs = []  # list of dataframes of F_Inputs sheets
foutputs_dfs = []  # list of dataframes of F_Outputs sheets
finputs_timestamps = [] # list of strings of F_Inputs timestamps
foutput_timestamps = [] # list of strings of CLEAR_SHEET output timestamp
finputs_runs = []  # list of strings of the CLEAR_SHEET run IDs
foutputs_runs = []  # list of strings of the CLEAR_SHEET run IDs
finputs_reportids = []  # list of strings of the CLEAR_SHEET F_inputs run IDs
foutputs_reportids = []  # list of strings of the CLEAR_SHEET F_Outputs run IDs
companyids = []  # list of strings of the CLEAR_SHEET F_Inputs company IDs


# temporary list of common names. In practice will be stored centrally somewhere
inputsheet = 'F_Inputs'
outputsheet = 'F_Outputs'
clearsheet = 'CLEAR_SHEET'

clearsheet_items = {'finputid':'F_Inputs_Report_ID', 'foutputid': 'F_Outputs_Report_ID',
                    'finputrun':'inputRunId', 'foutputrun': 'outputRunId',
                    'foutput_timestamp': 'outputSheetLastSent', 'companyid': 'companyId',
                    'tagid': 'tagId'}

QA_codes = {'QA1':{'QA_pattern' : '*_OUT1', 'QA_code':''},
            'QA2': {'QA_pattern' : '*_OUT2', 'QA_code':''}, 
            'QA3': {'QA_pattern' : '*_OUT3', 'QA_code':''},
            'QA4': {'QA_pattern' : '*_OUT4', 'QA_code':''}}


# The try except code needs to be replaced with more specific ways of capturing if sheet isn't present
def scrape_model(xlfile_name, inputsheet, outputsheet, clearsheet, clearsheet_items, QA_codes):
    try:
        dffinput = pd.read_excel(xlfile_name, sheet_name=inputsheet)  #  Create a dataframe from the F_Input sheet
    except:
        dffinput = pd.DataFrame() # create a blank dataframe if something goes wrong in the reading
        
    try:
        finput = dffinput.rename(columns=dffinput.iloc[0]).drop(dffinput.index[0]) # drop the extra index column the API provides
        finput = finput.iloc[1:] 
        finput = finput.Reference.tolist()  # Get a list of the BON codes
        finput = [x for x in finput if str(x) !='nan']  # Drop NANs from the list of BON codes
    except:
        finput = []
        
    try:
        finput_timestamp = dffinput.columns[4] # finds F_Inputs timestamp as the column header of the dataframe
    except:
        finput_timestamp = '' # blank string if F_INputs sheet isn't present
     
    try:
        dfoutput = pd.read_excel(file, sheet_name=outputsheet)  # Create a dataframe from the F_Output sheet
    except:
        dfoutput = pd.DataFrame() # create a blank dataframe if something goes wrong in the reading
        
    try:
        foutput = dfoutput.rename(columns=dfoutput.iloc[0]).drop(dfoutput.index[0])
        foutput = foutput.iloc[1:]
        foutput = foutput.Reference.tolist()   # Get the list of BON codes
        foutput = [x for x in foutput if str(x) !='nan']  # Drop NANs from the list of BON codes
    except:
        foutput = []
        
    # TODO Add if statement to skip if no CLEAR SHEET
    # read clearsheet items passed and create a dataframe
    clearsheet_metadata = pd.DataFrame.from_dict(clearsheet_items, orient = 'index', columns =['fields'])
    
    if clearsheet in pd.ExcelFile(xlfile_name).sheet_names:
        dfclearsheet = pd.read_excel(file, sheet_name=clearsheet, index_col=0, header= None, names=['metadata'])  #  Create a dataframe from the clearsheet
            
#  TODO: Comment - could convert dfclearsheet to dictionary using e.g. df.to_dict('records')  ??
        clearsheet_metadata = clearsheet_metadata.merge(dfclearsheet, left_on='fields', right_index=True, how = 'left') #m
        clearsheet_metadata.fillna('', inplace=True) # replaces nans from a failed merge with empty strings
    else: 
        print("clear sheet not found in", file)
        clearsheet_metadata['metadata'] = ''
        
    for QA in list(QA_codes.keys()): # Loops through QA codes in lst of F_Outputs codes
        pattern = QA_codes[QA]['QA_pattern'] # pulls out the pattern from the QA_code dict
        matching = fnmatch.filter(foutput,pattern) # creates a list of all QA_codes that match the pattern
        if len(matching) >0:  # if QA codes have been found, use the first one (QA codes in the list should be identical)
            QA_codes[QA]['QA_code'] = matching[0]
        else: QA_codes[QA]['QA_code'] = '' # if no QA codes have been found return blank string
                
           
    return finput, foutput, dffinput, dfoutput, finput_timestamp, clearsheet_metadata, QA_codes

    

for file in xlsxfiles:  # For each model file we are working with...
    finput, foutput, dffinput, dfoutput, finput_timestamp, clearsheet_metadata, QA_codes = scrape_model(file, inputsheet, outputsheet, clearsheet, clearsheet_items, QA_codes)  # ...get the list of BON codes in the f_input and f_output sheets
    finputs_codes.append(finput)  # Add the list of bon codes. This is creating a list of lists.
    foutputs_codes.append(foutput)  # Add the list of bon codes. This is creating a list of lists.
    finputs_dfs.append(dffinput) # Add to the list of F_Inputs dataframes
    foutputs_dfs.append(dfoutput) # Add to the list of F_Outputs dataframes
    finputs_timestamps.append(finput_timestamp) #Add straing timestamp to list of timestamps

    finputs_runs.append(clearsheet_metadata.loc['finputrun']['metadata'])  # Add to list of the CLEAR_SHEET run IDs
    foutputs_runs.append(clearsheet_metadata.loc['foutputrun']['metadata'])  # Add to list of the CLEAR_SHEET run IDs
    finputs_reportids.append(clearsheet_metadata.loc['finputid']['metadata'])  # Add to list of the CLEAR_SHEET F_inputs run IDs
    foutputs_reportids.append(clearsheet_metadata.loc['foutputid']['metadata'])  # Add to list of the CLEAR_SHEET F_Outputs run IDs
    companyids.append(clearsheet_metadata.loc['companyid']['metadata'])  # Add to list of the CLEAR_SHEET F_Inputs company IDs
    #QA1s = [] 
    #QA2s = []
    #QA3s = []
    

    
# create digraph, nodes and edges
num_models = len(xlsxfiles)
G = nx.DiGraph()  # Create an empty digraph with no nodes and no edges.

# Add attribtues for each model called F_Inputs and F_Outputs
for i in range(num_models):  # For each of the model files...
    G.add_node(xlsxfiles[i],  # make each model a node...
               F_Inputs=finputs_codes[i],  # ...associate it with its list of F_Input BON codes and ...
               F_Outputs=foutputs_codes[i])  # ...associate it with its list of F_Input BON codes

for i in range(num_models):  # For each of the model files...
    for j in range(num_models):  # For each of the model files...
        data_trans = [k for k in finputs_codes[i] if k in foutputs_codes[j]]
        if data_trans:
            # Add edge between node j and i, associate with BON codes
            G.add_edge(xlsxfiles[j], xlsxfiles[i], Data_transfered=data_trans)

all_F_inputs = sum(finputs_codes, []) #  This is a way to flatten the list of lists
BONs_nowhere = {}  # This dictionary will hold the models and its f_Output bons that go nowhere
for i in range(num_models):  # For each of the model files...
    data_not_trans = list(set(foutputs_codes[i]) - set(all_F_inputs))  # The data not transferred are the f_output BONs in model i that are not in the flat list of all f_input BONS
    BONs_nowhere.update({xlsxfiles[i]: data_not_trans})  # Update the dictionary

# Export the results (orient is required to overcome the need for each column (model) to have the same number of entries. It is transposed back later
pd.DataFrame.from_dict(BONs_nowhere, orient='index').T.to_excel("Outputs/BONS_nowhere.xlsx", index=False)


# Draw network
nx.draw_networkx(G)
plt.show()

# Check for cycles
try:
    cycles = nx.find_cycle(G)
    print("Cycles found")
    print(*cycles, sep="/n")
except nx.exception.NetworkXNoCycle:
    print("No cycles found")
    
# Work out batch order
'''
def find_order(G):
    nodes = list(G.nodes)
    for node in nodes:
'''

#  Export results
nx.write_gpickle(G, 'graph2.pkl')
pickle.dump(G, open('graph.pkl', 'wb'))
#nx.write_gml(G, "graph.graphml")
#nx.write_graphml(G, "graph.graphml")


end = time.time()
print("Time: ", round(end-start, 1))

