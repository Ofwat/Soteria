# -*- coding: utf-8 -*-
"""
Created on Wed Jan 18 14:16:11 2023

@author: h_jet
"""

from tkinter import Tk
from tkinter.filedialog import askdirectory
import os
import pandas as pd
import glob
import networkx as nx
import xlrd
from openpyxl import load_workbook
from pyxlsb import open_workbook

import time

start = time.time()

inputsheet = 'F_Inputs'
outputsheet = 'F_Outputs'
clearsheet = 'CLEAR_sheet'

path = askdirectory(title='Select Folder') # shows dialog box and return the path
print(path)

os.chdir(path)

# create list of all xls* files in root folder and subfolders and adds to xslxfiles if an "F_Inputs" or "F_Outputs" sheet is present
xlsxfiles = []
for file in glob.glob("**/*.xls*", recursive=True):
    filename, file_extension = os.path.splitext(file) 
    # check if files contain an F_INputs or F_Outputs sheet
    try: 
        if file_extension == '.xls':
            filexl = xlrd.open_workbook(file)
            if 'F_Inputs' in filexl.sheet_names() or 'F_Outputs' in filexl.sheet_names():
                xlsxfiles.append(file)
    
        elif file_extension == '.xlsb':
            filexl = open_workbook(file)
            if 'F_Inputs' in filexl.sheets or 'F_Outputs' in filexl.sheets:
                xlsxfiles.append(file)
            
        else: 
            filexl = load_workbook(file, read_only = True)
            if 'F_Inputs' in filexl.sheetnames or 'F_Outputs' in filexl.sheetnames:
                xlsxfiles.append(file)
                    
    except:
        print("error with", file)

# create lists to capture node attributes for all models

finputs_codes = [] # list of lists of F_Inputs boncodes
foutputs_codes = [] # list of lists of F_Outputs boncodes
finputs_dfs = [] # list of dataframes of F_Inputs sheets
foutputs_dfs = [] # list of dataframes of F_Outputs sheets
finputs_timestamps = [] # list of strings of F_Inputs timestamps
finputs_runs = [] # list of strings of the CLEAR_SHEET run IDs
foutputs_runs = [] # list of strings of the CLEAR_SHEET run IDs
finputs_reportids = [] # list of strings of the CLEAR_SHEET F_inputs run IDs
foutputs_reportids = [] # list of strings of the CLEAR_SHEET F_Outputs run IDs
finputs_companyids = [] # list of strings of the CLEAR_SHEET F_Inputs company IDs
foutputs_companyids = [] # list of strings of the CLEAR_SHEET F_Outputs company IDs
QA1s = []
QA2s = []
QA3s = []



    
def scrape_model(xlfile_name, inputsheet, outputsheet, clearsheet):
    try:
        dfinput = pd.read_excel(file, sheet_name=inputsheet)
        dfinput = dfinput.rename(columns=dfinput.iloc[0]).drop(dfinput.index[0])
        dfinput = dfinput.iloc[1:]
        finput = dfinput.Reference.tolist()
        finput = [x for x in finput if str(x) !='nan']
    except ValueError:
        finput = []
     
    try:
        dfoutput = pd.read_excel(file, sheet_name=outputsheet)
        dfoutput = dfoutput.rename(columns=dfoutput.iloc[0]).drop(dfoutput.index[0])
        dfoutput = dfoutput.iloc[1:]
        foutput = dfoutput.Reference.tolist()
        foutput = [x for x in finput if str(x) !='nan']
    except ValueError:
        foutput = []
        
    try:
        dfoutput = pd.read_excel(file, sheet_name=clearsheet)
        
    except:
        finputid = ''
            
    return finput, foutput

for file in xlsxfiles:
    
    finput, foutput = scrape_model(file, inputsheet, outputsheet, clearsheet)        
    finputs_codes.append(finput)
    foutputs_codes.append(foutput)
# create digraph, nodes and edges
num_models = len(xlsxfiles)
G = nx.DiGraph()
for i in range(num_models): G.add_node(xlsxfiles[i], F_Inputs = finputs_codes[i], F_Outputs = foutputs_codes[i])

for i in range(num_models): 
    for j in range(num_models):
        data_trans = [k for k in finputs_codes[i] if k in foutputs_codes[j]]
        if data_trans:
            G.add_edge(xlsxfiles[j], xlsxfiles[i], Data_transfered = data_trans)
        
# draw network
nx.draw_networkx(G)

# check for cycles
try:
    cycles = nx.find_cycle(G)
    print("cycles found")
    print(*cycles, sep="/n")
except nx.exception.NetworkXNoCycle:
    print("no cycles found")
    
# work out batch order
'''
def find_order(G):
    nodes = list(G.nodes)
    for node in nodes:
'''      

end = time.time()
print (end-start)

